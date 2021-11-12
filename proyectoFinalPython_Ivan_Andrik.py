#Billetera de Criptomonedas con CoinMarket
from datetime import datetime # importa la librería para obtener la fecha
# from openpyxl.descriptors.base import DateTime
import requests #importa requests previamente instalando requests desde pip
import openpyxl #importa para trabajar con archivos xslx

# Inicializa variables
monedas_dict={} #diccionario de criptomonedas
cotizacion={} #diccionario de los precios de criptomonedas
caracteresCodigo="abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789"
codigoPropio="ab123"

def validarCodigo(): #Procedimiento que comprueba y valida el código de Transacción
    a=set(caracteresCodigo)
    codigoIngresado=input("Ingrese el código: ")
    b=set(codigoIngresado)
    #Evalua y hace una diferencia entre set b y a para comprobar que el código cumpla
    # con el requisito de pertenecer a los caracteres permitidos
    while ((len(codigoIngresado)<4) or (len(b-a)>0) or (codigoIngresado==codigoPropio)):
        codigoIngresado=input("Código inválido. Ingrese el código: ")
        b=set(codigoIngresado)
    print("Código Válido")

# Archivos - billetera.xlsx (balance) Archivos/billetera.xlsx
direccionBilletera="Archivos/billetera.xlsx"

def cargarBilletera(): #abre el archivo para obtener datos
    libro=openpyxl.load_workbook(direccionBilletera)
    hoja=libro.worksheets[0] # asigna la hoja de calculo a hoja
    dictBilletera={} # Inicializa un diccionario
    for row in hoja.rows:
        if row[0].value!="Criptomoneda":
            dictBilletera[row[0].value]=row[1].value #Asigna a un diccionario los datos
    return dictBilletera

# Archivos - billetera.xlsx (Enviar-Recibir)
def cargarBilleteraEnvRec(): #abre el archivo para obtener datos
    libro=openpyxl.load_workbook(direccionBilletera)
    # retorna el libro y la hoja de calculo
    return libro,libro.worksheets[0]

# Archivos - transacciones.xlsx (historial) Archivos/transacciones.xlsx
direccionTransacciones="Archivos/transacciones.xlsx"

def cargarTransacciones(): #abre el archivo para obtener datos
    libro=openpyxl.load_workbook(direccionTransacciones)
    return libro,libro.worksheets[0] # retorna el libro y la hoja de calculo

def inicializacion(): #funcion que retorna el json data con lista de  Criptos
    # header para la conexión con la API de CoinMarket
    headers = {'Accepts': 'application/json', 'X-CMC_PRO_API_KEY': "d6e59a47-5206-46c5-8438-d7cf2d26ddcd"}
    # Solicitud del listado de CriptoMonedas obteniendo un json
    data=requests.get("https://pro-api.coinmarketcap.com/v1/cryptocurrency/listings/latest",headers=headers).json()
    return data

def listaCriptos(data): #función con parámetro que guarda Nombre y Precios para cada Criptomoneda
    for cripto in data["data"]:
        #Agrega el Symbol de cada cripto como key y el Nombre como dato al diccionario
        monedas_dict[cripto["symbol"]]=cripto["name"]
        #Agrega el Symbol de cada cripto como key y el Precio como dato al diccionario
        # y Redonea el precio a 2 decimales de USD
        cotizacion[cripto["symbol"]]=float("{:.2f}".format(round(cripto["quote"]["USD"]["price"],2)))

def validarCripto(lista): # Comprueba si la cripto ingresada es válida o no
    moneda=input("Indique el Símbolo de la moneda (Ej: BTC, ETH, etc): ")
    # se comprueba si es válida, sino solicita nuevamente
    while not (moneda in lista and moneda in monedas_dict.keys()): 
            moneda=input("Moneda Inválida. Ingrese una moneda válida: ")
    else:
        print("Eligió:",moneda,"-",monedas_dict[moneda],
            "(Verificado con coimnmarketcap.com)") # Moneda Verificada
    return moneda

def validarCriptoRecibir(): # Comprueba si la cripto ingresada es válida o no
    moneda=input("Indique el Símbolo de la moneda (Ej: BTC, ETH, etc): ")
    # se comprueba si es válida, sino solicita nuevamente
    while not (moneda in monedas_dict.keys()): 
            moneda=input("Moneda Inválida. Ingrese una moneda válida: ")
    else:
        print("Eligió:",moneda,"-",monedas_dict[moneda],
            "(Verificado con coimnmarketcap.com)") # Moneda Verificada
    return moneda

# Menú de Opciones
def menu():
    # Diccionario de Opciones
    opciones={"1":"Recibir Criptomoneda","2":"Transferir Criptomoneda",
    "3":"Mostrar balance de una Criptomoneda","4":"Mostrar balance general",
    "5":"Mostrar histórico de transacciones","6":"Salir del programa"}
    # Bienvenida y elección de opciones
    print("#######################")
    print("Billetera de Criptomonedas\n  Opciones:")
    for clave,valor in opciones.items(): #Imprime las Opciones
        print(clave,":",valor)
    irA=input("Elija una opción (número): ")
    while irA not in opciones: # comprueba si es válida, sino solicita nuevamente
        irA=input("Opción inválida, indique una válida: ")
    return irA

def irAlaOpcion(opcion): # Dirige a la opción elegida
    if opcion=="1":
        recibirCantidad() # Función para recibir una cantidad de una Criptomoneda
    elif opcion=="2":
        enviarCantidad() # Función para enviar una cantidad de una Criptomoneda
    elif opcion=="3":
        balance() # Muestra la sección del Balance
    elif opcion=="4":
        balanceGeneral() # Muestra el Balance General
    elif opcion=="5":
        historialTransacciones() # Muestra el historial de transacciones
    elif opcion=="6":
        print("Salir") # Cierra el programa

# Opción 1
def recibirCantidad():
    libroBill,billetera=cargarBilleteraEnvRec() #Obtengo los datos de la billetera
    billeteraDict={}
    for row in billetera.rows:
        if row[0].value!="Criptomoneda":
            billeteraDict[row[0].value]=row[1].value #Asigna a un diccionario los datos
    libroTra,transacciones=cargarTransacciones() # Obtengo el historial de transacciones
    print("#######################\nRecibir una Criptomoneda: ")
    listaCriptos(inicializacion()) # se obtiene la lista de criptomonedas desde Coinmarket
    moneda=validarCriptoRecibir() #se solicita y valida la criptomoneda elegida
    # Se muestra los datos de la cripto elegida
    print(moneda,"-", monedas_dict[moneda] ,": Cotización en USD= ",cotizacion[moneda])
    if (moneda in billeteraDict.keys()): #evalua si ya existe en su billetera
        print("     Su Saldo",billeteraDict[moneda])
        print("     Su Monto en USD:", #redondea y hace el cálculo en USD
        "{:.2f}".format(round(float(billeteraDict[moneda])*float(cotizacion[moneda]),2)))
        monto=float(input("Ingrese el monto a Recibir: "))
        validarCodigo()
        for row in billetera.rows: #busca en la billetera la criptomoneda elegida
            if row[0].value==moneda: #busca la criptomoneda en la billetera 
                row[1].value=float(row[1].value)+float(monto) #suma al saldo el monto indicado
                print("     Su Saldo final",row[1].value)
                print("     Su Monto en USD final",float(row[1].value)*float(cotizacion[moneda]))
                break #cuando encuentra, finaliza el for
        libroBill.save(filename=direccionBilletera)#guarda el libro Billetera.xlsx (balance)
        transacciones.append([(datetime.now()).strftime("%""d/%m/%Y"),moneda,monto,"{:.2f}".format(round(float(monto)*float(cotizacion[moneda]),2))])
        libroTra.save(filename=direccionTransacciones)
        print("Se recibió con éxito.")
    else:
        print("Usted aún no posee esta Criptomoneda")
        monto=float(input("Ingrese el monto a Recibir (NO USD): "))
        validarCodigo() #solicita y valida el código
        billetera.append([moneda,monto]) #añade al final una fila con la criptomoneda y el monto
        libroBill.save(filename=direccionBilletera)#guarda el libro Billetera.xlsx (historial)
        #asigna los valores a cada campo del libro Transacciones y los añade al libro
        transacciones.append([(datetime.now()).strftime("%""d/%m/%Y"),moneda,monto,"{:.2f}".format(round(float(monto)*float(cotizacion[moneda]),2))])
        # Guarda el libro Transacciones
        libroTra.save(filename=direccionTransacciones)
        print("Se recibió con éxito.")
    irAlaOpcion(menu()) #vuelve al menú

# Opción 2
def enviarCantidad():
    libroBill,billetera=cargarBilleteraEnvRec() #Obtengo los datos de la billetera
    billeteraDict={}
    for row in billetera.rows:
        if row[0].value!="Criptomoneda":
            billeteraDict[row[0].value]=row[1].value #Asigna a un diccionario los datos
    libroTra,transacciones=cargarTransacciones() # Obtengo el historial de transacciones
    print("#######################\nEnviar una Criptomoneda: ")
    print("Criptomoneda a Enviar: ")
    listaCriptos(inicializacion()) # se obtiene la lista de criptomonedas desde Coinmarket
    print("Las Criptomonedas de su Billetera: ")
    for clave,valor in billeteraDict.items():
        print(clave) # Muestra las Criptomonedas existentes en su billetera
    moneda=validarCripto(billeteraDict.keys()) # Se solicita y valida la Cripto ingresada
    # Se muestra los datos de la cripto elegida
    print(moneda,"-", monedas_dict[moneda] ,": Cotización en USD= ",cotizacion[moneda])
    if (moneda in billeteraDict.keys()): #evalua si ya existe en su billetera
        print("     Su Saldo",billeteraDict[moneda])
        print("     Su Monto en USD:", #redondea y hace el cálculo en USD
        "{:.2f}".format(round(float(billeteraDict[moneda])*float(cotizacion[moneda]),2)))
        monto=float(input("Ingrese el monto en USD a Enviar: ")) #solicita el ingreso del monto
        while float(monto) > (float(billeteraDict[moneda])*float(cotizacion[moneda])): #verifica que sea menor o igual al saldo
            monto=float(input("Monto inválido. Ingrese un monto menor o igual a su Saldo en USD: "))
        validarCodigo() #solicita y valida el código
        for row in billetera.rows: #busca en la billetera la criptomoneda elegida
            if row[0].value==moneda: #busca la criptomoneda en la billetera 
                #resta al saldo de la criptomoneda el monto indicado
                row[1].value=float(row[1].value)-(float(monto)/(float(cotizacion[moneda])))
                print("     Su Saldo final",row[1].value)
                print("     Su Monto en USD final",float(row[1].value)*float(cotizacion[moneda]))
                break #cuando encuentra, finaliza el for
        # Guarda el libro Billetera
        libroBill.save(filename=direccionBilletera)
        #asigna los valores a cada campo del libro Transacciones y los añade al libro
        transacciones.append([(datetime.now()).strftime("%""d/%m/%Y"),moneda,0-(float(monto)/(float(cotizacion[moneda]))),"{:.2f}".format(round(float(monto),2))])
        # Guarda el libro Transacciones
        libroTra.save(filename=direccionTransacciones)
        print("Se envió con éxito.")
    else:
        print("Usted aún no posee esta Criptomoneda.\nNo puede hacer un Envío.")
    irAlaOpcion(menu()) #vuelve al menú

# Opción 3
def balance(): # muestra el balance de una criptomoneda
    billetera=cargarBilletera() #Obtengo los datos de la billetera
    print("#######################\n: Balance de una Criptomoneda")
    print("Las Criptomonedas de su Billetera: ")
    for clave,valor in billetera.items():
        print(clave) # Muestra las Criptomonedas existentes en su billetera
    listaCriptos(inicializacion()) # obtiene las criptomonedas de coinmarket
    print("Elija una de la lista")
    moneda=validarCripto(billetera.keys()) # solicita el ingreso de una Cripto y valida
    # Se muestra los datos de la cripto elegida
    print(moneda,"-", monedas_dict[moneda] ,": Cotización en USD= ",cotizacion[moneda])
    print("     Saldo",billetera[moneda])
    print("     Monto en USD:", #redondea y hace el cálculo en USD
    "{:.2f}".format(round(float(billetera[moneda])*float(cotizacion[moneda]),2)))
    irAlaOpcion(menu()) #vuelve al menú

# Opción 4
def balanceGeneral(): # muestra el balance General
    billetera=cargarBilletera() #Obtengo los datos de la billetera
    listaCriptos(inicializacion()) # obtengo las criptomonedas de coinmarket
    print("#######################\nSu balance General: ")
    total=0.0 # acumula los Montos en USD
    # Se muestra los datos de las cripto de la billetera
    for clave,valor in billetera.items():
        print(clave,"-",monedas_dict[clave],": Cotización en USD= ",cotizacion[clave])
        print("     Saldo",valor)
        print("     Monto en USD:", #redondea y hace el cálculo en USD
        "{:.2f}".format(round(float(valor)*float(cotizacion[clave]),2)))
        total=total+float(valor)*float(cotizacion[clave]) #acumula el total en USD
    print("Monto Total en USD= ", "{:.2f}".format(round(float(total),2)))
    irAlaOpcion(menu()) #vuelve al menú

# Opción 5
def historialTransacciones(): # Muestra el Historial de Transacciones
    libroTra,transacciones=cargarTransacciones() # Abre el archivo de Transacciones
    print("#######################\nHistorial de Transacciones:")

    for row in transacciones.rows: # Muestra los datos del archivo
        if row[0].value=="Fecha":
            print(row[0].value,"    ",row[1].value,"    ",row[2].value,"    ",row[3].value)
        else:
            print(row[0].value,"    ",row[1].value,"            ",row[2].value,"        ",row[3].value)
    irAlaOpcion(menu())   #vuelve al menú

# ----------------- Inicio del programa ----------------- #

irAlaOpcion(menu()) # Se muestra el menú, se valida y redirige a la opción elegida
